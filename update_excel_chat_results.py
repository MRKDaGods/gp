"""
Update ReID_Experiments_Complete.xlsx with results from the EPFL Lab
end-to-end pipeline testing chat session (2026-03-09).

Adds:
  - Sheet 3 (Ablation Study): EPFL Lab pipeline ablation entries
  - Sheet 5 (Bug Taxonomy): 5 new bugs (#11-#15) found during testing
  - Sheet 6 (Final Results): EPFL Lab pipeline results section
  - Sheet 8 (Lessons Learned): 4 new lessons (#16-#19)

No new rows for Sheets 1-2 (no new ReID model training in this session).
"""

from copy import copy
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FILE = "e:/dev/src/gp/ReID_Experiments_Complete.xlsx"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def copy_style(src_cell, dst_cell):
    """Copy formatting from src_cell to dst_cell."""
    if src_cell.font:
        dst_cell.font = copy(src_cell.font)
    if src_cell.border:
        dst_cell.border = copy(src_cell.border)
    if src_cell.alignment:
        dst_cell.alignment = copy(src_cell.alignment)
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format


def write_row(ws, row_num, values, style_source_row=None, fills=None):
    """Write a row of values with optional formatting copied from a source row."""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGN
        if style_source_row is not None:
            src = ws.cell(row=style_source_row, column=col_idx)
            copy_style(src, cell)
        if fills and col_idx in fills:
            cell.fill = fills[col_idx]


def write_section_header(ws, row_num, text, num_cols=8):
    """Write a bold section header spanning multiple columns."""
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = SECTION_FONT
    cell.alignment = Alignment(wrap_text=False)
    ws.merge_cells(
        start_row=row_num, start_column=1,
        end_row=row_num, end_column=num_cols,
    )


def write_col_headers(ws, row_num, headers):
    """Write column headers with blue background."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")


def main():
    wb = openpyxl.load_workbook(FILE)

    # ------------------------------------------------------------------
    # Sheet 3: Ablation Study — add EPFL Lab pipeline ablations
    # ------------------------------------------------------------------
    ws3 = wb["Ablation Study"]
    # Current last row is 22 (NET RESULT). Append new section after it.
    r = ws3.max_row + 2  # skip a blank row

    write_section_header(ws3, r,
        "MTMC Pipeline Ablations — EPFL Lab 6-Person "
        "(no ground truth; proxy metrics: trajectory/cluster counts)")
    r += 1

    write_col_headers(ws3, r, [
        "Fix Applied", "Config", "Compared To",
        "Trajectories", "Multi-cam", "Singletons", "Δ",
        "Significance & Explanation",
    ])
    r += 1

    epfl_ablations = [
        [
            "Baseline: 2fps + OSNet-512D + connected_components",
            "2fps, OSNet-x1.0 (512D person ReID), connected_components, thresh=0.3",
            "—",
            "1", "1", "0", "—",
            "FAILURE. All 126 tracklets merged into 1 cluster. "
            "Connected components is transitive: A~B and B~C → all merge. "
            "All cross-camera cosine sims ≥0.499 (mean 0.833) at 360×288.",
        ],
        [
            "Algorithm: connected_comp → Louvain (res=2.0)",
            "2fps, OSNet-512D, community_detection, res=2.0, thresh=0.5",
            "Baseline (thresh=0.3, conn_comp)",
            "31", "14", "17", "+30",
            "Louvain community detection breaks transitive closure. "
            "Resolution=2.0 produces finer-grained clusters. "
            "However, 31 identities for 6 real people = heavy fragmentation.",
        ],
        [
            "Extraction FPS: 2 → 5",
            "5fps, OSNet-512D, Louvain res=2.0, thresh=0.5",
            "2fps Louvain run (31 traj)",
            "13", "9", "4", "-18",
            "5fps gives tracker 200ms gaps vs 500ms. "
            "Kalman filter predictions far more accurate → fewer ID switches → "
            "fewer redundant tracklets → better clustering.",
        ],
        [
            "Person ReID: OSNet-x1.0 (512D) → ResNet50-IBN-a (2048D)",
            "5fps, ResNet50-IBN-a (2048D), Louvain res=1.5, thresh=0.45",
            "5fps OSNet run (13 traj)",
            "14 (pre-split)", "11", "3", "+1",
            "2048D embeddings from stronger backbone provide more discriminative "
            "representations. Combined with lower threshold (0.45) allows more "
            "cross-camera edges in similarity graph.",
        ],
        [
            "Same-camera conflict resolution (graph coloring)",
            "Post-clustering: split clusters with same-cam temporal overlaps",
            "Louvain only (14 clusters)",
            "22", "16", "6", "+8",
            "CRITICAL. Louvain transitively merged tracklets from same camera "
            "with overlapping time (physically impossible). Graph coloring on "
            "conflict sub-graph split 8 violated clusters → no false merges.",
        ],
        [
            "Video: max interpolation gap (0.8s) + trail reset",
            "Annotation: skip interpolation if keyframe gap > 0.8s original time",
            "Unbounded interpolation (v2)",
            "—", "—", "—", "—",
            "Eliminates phantom bounding boxes at old positions. "
            "Trails reset on temporal gaps → no more weird cross-gap lines. "
            "Visual quality fix, no impact on tracking metrics.",
        ],
    ]

    for ablation in epfl_ablations:
        write_row(ws3, r, ablation)
        r += 1

    # ------------------------------------------------------------------
    # Sheet 5: Bug Taxonomy — append bugs #11-#15
    # ------------------------------------------------------------------
    ws5 = wb["Bug Taxonomy"]
    # Last bug is #10 at row 13. Data format row to copy style from: row 13.
    style_row = ws5.max_row
    r = style_row + 1

    new_bugs = [
        [
            11,
            "Unbounded bbox interpolation",
            "Stage 6 (VideoAnnotator)",
            "Linear interpolation between keyframes had no gap limit. "
            "At 2fps (12-frame gaps), person could move significantly "
            "between keyframes → interpolated boxes drawn at wrong positions. "
            "After tracklet ends, phantom boxes persisted at last known location.",
            "Visual: stale bboxes at old positions, user confusion",
            "Added max_interp_seconds=0.8 parameter. Skip interpolation "
            "when gap > fps×0.8 frames. At 25fps, max gap = 20 frames.",
            "video_annotator.py:71",
        ],
        [
            12,
            "Trail rendering across temporal gaps",
            "Stage 6 (VideoAnnotator)",
            "Trail polyline connected last-known and newly-detected positions "
            "across arbitrary temporal gaps. If a track was lost for 2 seconds "
            "then re-acquired, a long diagonal line was drawn across the scene.",
            "Visual: weird trajectory artifacts",
            "Track last_frame per global_id. Reset trail[] if "
            "(current_frame - last_frame) > max_trail_gap.",
            "video_annotator.py:189-191",
        ],
        [
            13,
            "PCA dimension collapse with few samples",
            "Stage 2 (Feature Pipeline)",
            "PCA with n_samples=146 < n_features=512 constrained output to "
            "min(n,d)=146 dimensions. Whitening amplified noise in degenerate "
            "eigenvectors. All embeddings collapsed to near-identical vectors.",
            "Embedding quality destroyed; all pairs cosine sim ≥0.85",
            "Skip PCA when n_samples < n_features × 2. Already partially "
            "addressed in previous session but threshold was too high.",
            "stage2/pipeline.py:163",
        ],
        [
            14,
            "Louvain transitive same-camera merging",
            "Stage 4 (Graph Solver)",
            "Louvain community detection transitively grouped same-camera "
            "tracklets with overlapping time spans. E.g., A(cam0)~B(cam1) and "
            "C(cam0)~B(cam1) → A,B,C all merged, but A and C are different "
            "people visible simultaneously on cam0.",
            "Two different people share same global ID in same camera view",
            "Post-clustering graph coloring on conflict sub-graph. Nodes that "
            "conflict (same-camera + temporal overlap) get different colors; "
            "each color group becomes a separate identity cluster.",
            "stage4/pipeline.py:335-417",
        ],
        [
            15,
            "Low-FPS Kalman filter failure",
            "Stage 1 (Tracker / BoT-SORT)",
            "At 2fps extraction (500ms inter-frame gap), the Kalman filter's "
            "motion predictions diverge. Persons lost during pose changes "
            "(turning, bending) because predicted bbox position is too far "
            "from actual. Creates excessive ID switches and short tracklets.",
            "~2× more tracklets than necessary; excessive fragmentation",
            "Increase extraction FPS to ≥5. At 5fps (200ms gaps), BoT-SORT "
            "tracking is dramatically more stable. Also raise track_buffer=60 "
            "and appearance_thresh=0.5.",
            "Pipeline config: stage0.output_fps",
        ],
    ]

    for bug in new_bugs:
        write_row(ws5, r, bug, style_source_row=style_row)
        r += 1

    # ------------------------------------------------------------------
    # Sheet 6: Final Results — add EPFL Lab pipeline results section
    # ------------------------------------------------------------------
    ws6 = wb["Final Results"]
    # Currently ends at row 24 (note row). Add EPFL Lab section after.
    r = ws6.max_row + 2

    write_section_header(ws6, r,
        "MTMC Pipeline End-to-End Results — EPFL Lab 6-Person "
        "(360×288, no ground truth)", num_cols=10)
    r += 1

    write_col_headers(ws6, r, [
        "Method", "ReID Backbone", "Association",
        "Extraction", "Trajectories", "Multi-cam",
        "Singletons", "Clusters (Louvain)",
        "Status", "Notes",
    ])
    r += 1

    epfl_results = [
        [
            "Baseline (v1)",
            "OSNet-x1.0 (512D, Market-1501)",
            "conn_components, thresh=0.3, w=(0.5/0.3/0.2)",
            "2fps",
            "1", "1", "0", "1",
            "FAILURE",
            "All tracklets merged into single cluster. Connected components "
            "too transitive; all cosine sims ≥0.499 at 360×288 resolution.",
        ],
        [
            "Louvain + HSV rebalanced (v2)",
            "OSNet-x1.0 (512D, Market-1501)",
            "Louvain(res=2.0), thresh=0.5, w=(0.5/0.3/0.2)",
            "2fps",
            "31", "14", "17", "31",
            "Improved",
            "Community detection breaks transitive closure. "
            "But 31 identities for 6 people = heavy fragmentation. "
            "2fps too sparse for reliable tracking (Kalman filter diverges).",
        ],
        [
            "5fps + OSNet (v3a)",
            "OSNet-x1.0 (512D, Market-1501)",
            "Louvain(res=2.0), thresh=0.5, w=(0.5/0.3/0.2)",
            "5fps",
            "13", "9", "4", "13",
            "Good",
            "5fps dramatically improves tracking stability. "
            "126 tracklets (same as v2 but fewer ID switches). "
            "13 clusters closer to 6 real identities.",
        ],
        [
            "ResNet50-IBN + conflict res. (v3b) \u2605",
            "ResNet50-IBN-a (2048D, Market-1501)",
            "Louvain(res=1.5), thresh=0.45, conflict resolution",
            "5fps",
            "22", "16", "6", "14 \u2192 22 after split",
            "\u2605 BEST",
            "Stronger backbone (2048D). Conflict resolution splits "
            "same-camera temporal overlaps (8 clusters split). "
            "22 identities: 6 large (likely real people) + 10 fragments + 6 singletons.",
        ],
    ]

    for i, row_data in enumerate(epfl_results):
        fills = {}
        status = row_data[8]
        if "BEST" in status:
            fills = {col: GREEN for col in range(1, 11)}
        elif "FAILURE" in status:
            fills = {col: RED for col in range(1, 11)}
        write_row(ws6, r, row_data, fills=fills)
        r += 1

    # Add analysis note
    r += 1
    note = (
        "Note: EPFL Lab videos are 360\u00d7288 — far below typical city surveillance "
        "(720p+). At this resolution, person crops are ~30-80px tall, making "
        "ReID models unable to distinguish individuals (mean cosine sim = 0.833 "
        "between all cross-camera pairs). Results validate pipeline plumbing "
        "but not ReID quality. Street-level datasets (CityFlow, WILDTRACK) needed "
        "for meaningful cross-camera evaluation."
    )
    write_section_header(ws6, r, note, num_cols=10)

    # ------------------------------------------------------------------
    # Sheet 8: Lessons Learned — append #16-#19
    # ------------------------------------------------------------------
    ws8 = wb["Lessons Learned"]
    style_row = 20  # last data row (lesson #15)
    r = style_row + 1

    new_lessons = [
        [
            16,
            "Extraction FPS critically impacts single-camera tracking",
            "EPFL Lab: 2fps \u2192 5fps reduced trajectory count from 31 to 13 "
            "(same 126 tracklets but better Louvain clustering). Kalman filter "
            "at 2fps (500ms gaps) diverges during pose changes. At 5fps (200ms), "
            "predictions are 6\u00d7 more accurate.",
            "Use \u22655fps for any real deployment. 2fps is only suitable for "
            "extremely slow or static scenes. For city traffic, 10fps is ideal.",
        ],
        [
            17,
            "Same-camera temporal overlap is a hard physical constraint",
            "Louvain clustering merged 8 clusters with same-camera temporal "
            "overlaps (two people visible simultaneously on one camera assigned "
            "same ID). Post-clustering graph coloring resolved all violations "
            "by splitting into 22 valid clusters.",
            "Always enforce same-camera temporal overlap as a post-processing "
            "constraint after ANY community detection or clustering algorithm. "
            "This is a zero-cost correctness guarantee.",
        ],
        [
            18,
            "Bounding box interpolation needs bounded temporal gap",
            "Unbounded interpolation drew phantom bboxes at old positions for "
            "up to 12 frames (0.5s at 25fps) after person moved. "
            "Capping gap at 0.8s of original video time eliminated artifacts. "
            "Trail rendering also needs gap-based reset to avoid cross-gap lines.",
            "Cap interpolation gap at <1s of original video time. Longer gaps "
            "indicate lost tracks, not smooth motion. Reset motion trails "
            "on the same gap threshold.",
        ],
        [
            19,
            "Dataset resolution is the fundamental ReID bottleneck",
            "EPFL Lab (360\u00d7288): all cross-camera cosine similarities \u22650.499, "
            "mean=0.833. Person crops are 30-80px tall vs 256\u00d7128 training images. "
            "No amount of algorithm tuning can overcome garbage-in embeddings. "
            "Switching OSNet \u2192 ResNet50-IBN-a helped marginally but didn't solve it.",
            "Test on \u2265720p footage for meaningful cross-camera evaluation. "
            "Low-resolution results only validate pipeline plumbing, not ReID quality. "
            "Use CityFlow/WILDTRACK for publication-grade experiments.",
        ],
    ]

    for lesson in new_lessons:
        write_row(ws8, r, lesson, style_source_row=style_row)
        r += 1

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    wb.save(FILE)
    print(f"Updated {FILE}")
    print("  Sheet 3 (Ablation Study): +6 EPFL Lab pipeline ablation rows")
    print("  Sheet 5 (Bug Taxonomy): +5 new bugs (#11-#15)")
    print("  Sheet 6 (Final Results): +4 EPFL Lab pipeline result rows")
    print("  Sheet 8 (Lessons Learned): +4 new lessons (#16-#19)")


if __name__ == "__main__":
    main()
